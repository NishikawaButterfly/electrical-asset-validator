from __future__ import annotations

import json
from collections.abc import AsyncIterator
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from electrical_asset_validator.services.ingest import (
    CANONICAL_COLUMNS,
    MAX_CELL_CHARACTERS,
)
from tests.conftest import (
    FORMULA_RESULTS,
    FORMULA_ROWS,
    csv_bytes,
    with_cached_values,
    xlsx_bytes,
)

pytestmark = pytest.mark.anyio


def _files(content: bytes, filename: str = "assets.csv") -> dict[str, tuple[str, bytes, str]]:
    return {"file": (filename, content, "text/csv")}


async def test_health(client: AsyncClient) -> None:
    response = await client.get("/api/v1/health")

    assert response.status_code == 200
    assert response.json() == {
        "status": "ok",
        "version": "0.2.0",
        "database": "ok",
    }


async def test_validation_history_detail_and_reports(
    client: AsyncClient,
    clean_csv: bytes,
) -> None:
    created = await client.post("/api/v1/validations", files=_files(clean_csv))

    assert created.status_code == 201
    payload = created.json()
    assert set(payload) == {
        "id",
        "filename",
        "created_at",
        "quality_score",
        "metrics",
        "issues",
        "download_urls",
    }
    assert payload["filename"] == "assets.csv"
    assert payload["quality_score"] == 100
    assert payload["metrics"] == {
        "total_rows": 2,
        "valid_rows": 2,
        "issue_count": 0,
        "returned_issue_count": 0,
        "issues_truncated": False,
        "error_count": 0,
        "warning_count": 0,
        "info_count": 0,
    }
    assert payload["issues"] == []

    history = await client.get("/api/v1/validations")
    assert history.status_code == 200
    assert isinstance(history.json(), list)
    assert history.json()[0]["id"] == payload["id"]

    detail = await client.get(f"/api/v1/validations/{payload['id']}")
    assert detail.status_code == 200
    assert detail.json() == payload

    excel = await client.get(payload["download_urls"]["excel"])
    assert excel.status_code == 200
    assert excel.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(excel.content))
    assert workbook.sheetnames == ["Summary", "Issues", "Data"]
    assert workbook["Data"]["A2"].value == "PNL-A"
    assert workbook["Summary"]["B4"].value == payload["created_at"]

    pdf = await client.get(payload["download_urls"]["pdf"])
    assert pdf.status_code == 200
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert pdf.content.startswith(b"%PDF")


async def test_validation_issues_are_persisted_with_string_ids(
    client: AsyncClient,
    clean_rows: list[dict[str, object]],
) -> None:
    clean_rows[1]["power_kw"] = -5
    created = await client.post(
        "/api/v1/validations",
        files=_files(csv_bytes(clean_rows), "invalid.csv"),
    )

    assert created.status_code == 201
    payload = created.json()
    assert payload["metrics"]["error_count"] == 1
    assert payload["issues"][0]["rule"] == "VALUE_OUT_OF_RANGE"
    assert isinstance(payload["issues"][0]["id"], str)

    detail = await client.get(f"/api/v1/validations/{payload['id']}")
    assert detail.json()["issues"] == payload["issues"]


async def test_comparison_is_persisted_and_uses_source_row_numbers(
    client: AsyncClient,
    clean_rows: list[dict[str, object]],
) -> None:
    before_rows = [dict(row) for row in clean_rows]
    after_rows = [dict(row) for row in clean_rows]
    after_rows[1]["power_kw"] = 20
    after_rows.append(
        {
            "asset_tag": "FAN-001",
            "asset_name": "Ventilation Fan",
            "asset_type": "fan",
            "location": "Plant 1",
            "panel_tag": "PNL-A",
            "circuit_ref": "C-02",
            "voltage_v": 400,
            "power_kw": 5,
            "status": "standby",
        }
    )

    response = await client.post(
        "/api/v1/comparisons",
        files={
            "before_file": ("before.csv", csv_bytes(before_rows), "text/csv"),
            "after_file": ("after.csv", csv_bytes(after_rows), "text/csv"),
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["summary"] == {
        "added": 1,
        "removed": 0,
        "changed": 1,
        "unchanged": 1,
    }
    assert payload["added"] == [{"asset_tag": "FAN-001", "row": 4}]
    assert payload["changed"] == [
        {
            "asset_tag": "MTR-001",
            "changes": [{"field": "power_kw", "before": 15, "after": 20}],
        }
    ]

    detail = await client.get(f"/api/v1/comparisons/{payload['id']}")
    assert detail.status_code == 200
    assert detail.json() == payload


async def test_a_workbook_with_calculated_formulas_validates_on_its_values(
    client: AsyncClient,
) -> None:
    content = with_cached_values(xlsx_bytes(FORMULA_ROWS), FORMULA_RESULTS)

    response = await client.post("/api/v1/validations", files=_files(content, "assets.xlsx"))

    assert response.status_code == 201
    payload = response.json()
    assert payload["quality_score"] == 100
    assert payload["issues"] == []


async def test_uncalculated_formulas_are_refused_as_one_fact_about_the_file(
    client: AsyncClient,
) -> None:
    content = xlsx_bytes(FORMULA_ROWS)

    response = await client.post("/api/v1/validations", files=_files(content, "assets.xlsx"))

    # An upload the software cannot read faithfully is refused the way every
    # other unreadable upload is: one status, one sentence, nothing stored.
    assert response.status_code == 400
    assert response.json() == {
        "detail": (
            "The XLSX workbook has 3 formulas with no stored result, the first "
            "in cell G2. Open it in Excel or LibreOffice and save it so the "
            "results are stored, or replace the formulas with their values, "
            "then upload it again."
        )
    }
    history = await client.get("/api/v1/validations")
    assert history.json() == []


async def test_validation_maps_nonstandard_headers_with_a_full_mapping(
    client: AsyncClient,
) -> None:
    content = (
        b"Equipment ID,Description,Category,Site,Feeder Panel,Feeder Circuit,"
        b"Volts,Kilowatts,State\n"
        b"PNL-A,Main Panel,panel,Plant 1,,,400,0,active\n"
        b"MTR-001,Conveyor Motor,motor,Plant 1,PNL-A,C-01,400,15,active\n"
    )
    mapping = {
        "Equipment ID": "asset_tag",
        "Description": "asset_name",
        "Category": "asset_type",
        "Site": "location",
        "Feeder Panel": "panel_tag",
        "Feeder Circuit": "circuit_ref",
        "Volts": "voltage_v",
        "Kilowatts": "power_kw",
        "State": "status",
    }

    response = await client.post(
        "/api/v1/validations",
        files=_files(content, "renamed.csv"),
        data={"mapping": json.dumps(mapping)},
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["quality_score"] == 100
    assert payload["issues"] == []
    assert payload["metrics"]["total_rows"] == 2


async def test_validation_maps_renamed_columns_with_a_partial_mapping(
    client: AsyncClient,
    clean_csv: bytes,
) -> None:
    content = clean_csv.replace(b"asset_tag", b"Equipment ID", 1)

    response = await client.post(
        "/api/v1/validations",
        files=_files(content, "partial.csv"),
        data={"mapping": json.dumps({"Equipment ID": "asset_tag"})},
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["quality_score"] == 100
    assert payload["issues"] == []


@pytest.mark.parametrize(
    ("mapping", "message"),
    [
        ("{not json", "must be a valid JSON object"),
        ('{"Voltage V": "volts"}', "'volts' is not a canonical field"),
        (
            '{"Asset Tag": "asset_tag", "Asset Name": "asset_tag"}',
            "map to the same canonical field: asset_tag",
        ),
        (
            '{"Missing Header": "asset_tag"}',
            "not present in the file: missing_header",
        ),
    ],
)
async def test_invalid_column_mappings_return_422(
    client: AsyncClient,
    clean_csv: bytes,
    mapping: str,
    message: str,
) -> None:
    response = await client.post(
        "/api/v1/validations",
        files=_files(clean_csv),
        data={"mapping": mapping},
    )

    assert response.status_code == 422
    assert message in response.json()["detail"]


async def test_comparison_applies_one_mapping_to_both_files(
    client: AsyncClient,
    clean_rows: list[dict[str, object]],
) -> None:
    after_rows = [dict(row) for row in clean_rows]
    after_rows[1]["power_kw"] = 20
    before_csv = csv_bytes(clean_rows).replace(b"asset_tag", b"Equipment ID", 1)
    after_csv = csv_bytes(after_rows).replace(b"asset_tag", b"Equipment ID", 1)

    response = await client.post(
        "/api/v1/comparisons",
        files={
            "before_file": ("before.csv", before_csv, "text/csv"),
            "after_file": ("after.csv", after_csv, "text/csv"),
        },
        data={"mapping": json.dumps({"Equipment ID": "asset_tag"})},
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["summary"] == {
        "added": 0,
        "removed": 0,
        "changed": 1,
        "unchanged": 1,
    }
    assert payload["changed"] == [
        {
            "asset_tag": "MTR-001",
            "changes": [{"field": "power_kw", "before": 15, "after": 20}],
        }
    ]


async def test_invalid_uploads_and_missing_records_return_clear_errors(
    client: AsyncClient,
) -> None:
    unsupported = await client.post(
        "/api/v1/validations",
        files=_files(b"not a register", "assets.txt"),
    )
    assert unsupported.status_code == 400
    assert "Only .csv and .xlsx" in unsupported.json()["detail"]

    assert (await client.get("/api/v1/validations/not-found")).status_code == 404
    assert (await client.get("/api/v1/comparisons/not-found")).status_code == 404


async def test_validation_rejects_an_oversized_multipart_body(
    client: AsyncClient,
    clean_csv: bytes,
) -> None:
    response = await client.post(
        "/api/v1/validations",
        files={
            "file": ("assets.csv", clean_csv, "text/csv"),
            "ignored": (
                "ignored.bin",
                b"x" * (3 * 1024 * 1024),
                "application/octet-stream",
            ),
        },
    )

    assert response.status_code == 413
    assert "request body is too large" in response.json()["detail"]


@pytest.mark.parametrize("content_length", [None, "1", "not-a-number"])
async def test_streamed_multipart_overflow_returns_413(
    client: AsyncClient,
    clean_csv: bytes,
    content_length: str | None,
) -> None:
    boundary = "eav-stream-boundary"

    async def request_body() -> AsyncIterator[bytes]:
        yield (
            f"--{boundary}\r\n"
            'Content-Disposition: form-data; name="file"; filename="assets.csv"\r\n'
            "Content-Type: text/csv\r\n\r\n"
        ).encode()
        yield clean_csv
        yield (
            f"\r\n--{boundary}\r\n"
            'Content-Disposition: form-data; name="ignored"; filename="ignored.bin"\r\n'
            "Content-Type: application/octet-stream\r\n\r\n"
        ).encode()
        for _ in range(48):
            yield b"x" * (64 * 1024)
        yield f"\r\n--{boundary}--\r\n".encode()

    headers = {"Content-Type": f"multipart/form-data; boundary={boundary}"}
    if content_length is not None:
        headers["Content-Length"] = content_length

    response = await client.post(
        "/api/v1/validations",
        content=request_body(),
        headers=headers,
    )

    assert response.status_code == 413
    assert "request body is too large" in response.json()["detail"]


async def test_excel_report_neutralizes_user_derived_formula_cells(
    client: AsyncClient,
) -> None:
    content = (
        ",".join([*CANONICAL_COLUMNS, "=1+1"])
        + "\n"
        + "PNL-A,Main Panel,panel,Plant 1,,,400,0,active,fictional\n"
    ).encode()
    created = await client.post(
        "/api/v1/validations",
        files=_files(content, "formula.csv"),
    )

    assert created.status_code == 201
    report = await client.get(created.json()["download_urls"]["excel"])
    workbook = load_workbook(BytesIO(report.content), data_only=False)
    field_cell = workbook["Issues"]["E2"]

    assert field_cell.value == "=1+1"
    assert field_cell.data_type == "s"
    assert field_cell.quotePrefix is True


async def test_pdf_report_handles_the_longest_accepted_header(
    client: AsyncClient,
) -> None:
    long_header = "X" * MAX_CELL_CHARACTERS
    content = (
        ",".join([*CANONICAL_COLUMNS, long_header])
        + "\n"
        + "PNL-A,Main Panel,panel,Plant 1,,,400,0,active,fictional\n"
    ).encode()
    created = await client.post(
        "/api/v1/validations",
        files=_files(content, "long-header.csv"),
    )

    assert created.status_code == 201
    report = await client.get(created.json()["download_urls"]["pdf"])

    assert report.status_code == 200
    assert report.content.startswith(b"%PDF")


async def test_excel_report_preserves_a_max_length_formula_like_value(
    client: AsyncClient,
) -> None:
    formula_like_name = "=" + ("A" * (MAX_CELL_CHARACTERS - 1))
    content = (
        ",".join(CANONICAL_COLUMNS)
        + "\n"
        + f"PNL-A,{formula_like_name},panel,Plant 1,,,400,0,active\n"
    ).encode()
    created = await client.post(
        "/api/v1/validations",
        files=_files(content, "max-cell.csv"),
    )

    assert created.status_code == 201
    report = await client.get(created.json()["download_urls"]["excel"])
    workbook = load_workbook(BytesIO(report.content), data_only=False)
    name_cell = workbook["Data"]["B2"]

    assert name_cell.value == formula_like_name
    assert len(name_cell.value) == MAX_CELL_CHARACTERS
    assert name_cell.data_type == "s"
    assert name_cell.quotePrefix is True


async def test_inspection_matches_standard_headers(
    client: AsyncClient,
    clean_csv: bytes,
) -> None:
    response = await client.post("/api/v1/inspections", files=_files(clean_csv))

    assert response.status_code == 200
    payload = response.json()
    assert payload["filename"] == "assets.csv"
    assert payload["columns"] == [
        {"header": field, "canonical_field": field} for field in CANONICAL_COLUMNS
    ]
    assert payload["unmatched_canonical_fields"] == []


async def test_inspection_reports_unmatched_renamed_headers(
    client: AsyncClient,
) -> None:
    content = (
        b"Equipment ID,Asset Name,asset_type,location,panel_tag,circuit_ref,"
        b"Volts,power_kw,status\n"
        b"PNL-A,Main Panel,panel,Plant 1,,,400,0,active\n"
    )

    response = await client.post(
        "/api/v1/inspections",
        files=_files(content, "renamed.csv"),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["columns"][0] == {
        "header": "Equipment ID",
        "canonical_field": None,
    }
    assert payload["columns"][1] == {
        "header": "Asset Name",
        "canonical_field": "asset_name",
    }
    assert payload["columns"][6] == {"header": "Volts", "canonical_field": None}
    assert payload["unmatched_canonical_fields"] == ["asset_tag", "voltage_v"]


async def test_inspection_rejects_invalid_and_oversized_files(
    client: AsyncClient,
) -> None:
    unsupported = await client.post(
        "/api/v1/inspections",
        files=_files(b"not a register", "assets.txt"),
    )
    assert unsupported.status_code == 400
    assert "Only .csv and .xlsx" in unsupported.json()["detail"]

    oversized = await client.post(
        "/api/v1/inspections",
        files=_files(b"x" * (3 * 1024 * 1024)),
    )
    assert oversized.status_code == 413
    assert "request body is too large" in oversized.json()["detail"]
