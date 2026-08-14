from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

from electrical_asset_validator.services import ingest
from electrical_asset_validator.services.ingest import (
    CANONICAL_COLUMNS,
    DatasetError,
    parse_dataset,
)
from electrical_asset_validator.services.validation import validate_dataset
from tests.conftest import (
    FORMULA_RESULTS,
    FORMULA_ROWS,
    with_cached_values,
    xlsx_bytes,
)


def test_csv_headers_are_safely_normalized() -> None:
    content = (
        b"Asset Tag,Asset Name,Asset Type,Location,Panel Tag,Circuit Ref,"
        b"Voltage V,Power KW,Status\n"
        b"PNL-A,Main Panel,panel,Plant 1,,,400,0,active\n"
    )

    dataset = parse_dataset("assets.csv", content)

    assert dataset.present_columns == set(CANONICAL_COLUMNS)
    assert dataset.records[0]["asset_tag"] == "PNL-A"
    assert dataset.records[0]["panel_tag"] is None
    assert dataset.records[0]["voltage_v"] == 400
    assert dataset.records[0]["power_kw"] == 0


def test_xlsx_first_worksheet_is_supported() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(CANONICAL_COLUMNS)
    sheet.append(["PNL-A", "Main Panel", "panel", "Plant 1", None, None, 400, 0, "active"])
    output = BytesIO()
    workbook.save(output)

    dataset = parse_dataset("assets.xlsx", output.getvalue())

    assert len(dataset.records) == 1
    assert dataset.records[0]["voltage_v"] == 400


def test_xlsx_formulas_are_read_as_the_result_the_workbook_stores() -> None:
    content = with_cached_values(xlsx_bytes(FORMULA_ROWS), FORMULA_RESULTS)

    dataset = parse_dataset("assets.xlsx", content)

    assert [record["voltage_v"] for record in dataset.records] == [400, 400]
    assert [record["power_kw"] for record in dataset.records] == [0, 15]
    # The register is sound, so it must validate as one rather than as a wall
    # of invalid-number findings.
    assert validate_dataset(dataset).quality_score == 100


def test_xlsx_formulas_with_no_stored_result_are_refused_once_for_the_file() -> None:
    content = xlsx_bytes(FORMULA_ROWS)

    with pytest.raises(DatasetError) as excinfo:
        parse_dataset("assets.xlsx", content)

    assert str(excinfo.value) == (
        "The XLSX workbook has 3 formulas with no stored result, the first in "
        "cell G2. Open it in Excel or LibreOffice and save it so the results "
        "are stored, or replace the formulas with their values, then upload it "
        "again."
    )


def test_a_single_uncalculated_formula_is_named_by_its_cell() -> None:
    content = xlsx_bytes(FORMULA_ROWS)
    # Every formula but one now carries the result a spreadsheet would cache.
    content = with_cached_values(content, {"G2": "400", "G3": "400"})

    with pytest.raises(DatasetError) as excinfo:
        parse_dataset("assets.xlsx", content)

    assert str(excinfo.value).startswith(
        "The XLSX workbook has a formula in cell H3 with no stored result."
    )


def test_text_that_merely_looks_like_a_formula_is_left_alone() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(CANONICAL_COLUMNS)
    sheet.append(["PNL-A", "Main Panel", "panel", "Plant 1", None, None, 400, 0, "active"])
    name_cell = sheet.cell(row=2, column=2)
    name_cell.value = "=200*2"
    name_cell.data_type = "s"
    output = BytesIO()
    workbook.save(output)

    dataset = parse_dataset("assets.xlsx", output.getvalue())

    # A text cell is not a formula, so it is neither resolved nor refused.
    assert dataset.records[0]["asset_name"] == "=200*2"


def test_a_formula_the_workbook_stored_as_an_error_is_reported_by_the_rules() -> None:
    content = xlsx_bytes(FORMULA_ROWS)
    content = with_cached_values(content, {"G2": "400", "G3": "400", "H3": "#DIV/0!"})

    dataset = parse_dataset("assets.xlsx", content)

    # The workbook does hold a result; it is the spreadsheet's own error value.
    # That belongs to the row, not to the file, so the numeric rule reports it.
    assert dataset.records[1]["power_kw"] == "#DIV/0!"
    findings = validate_dataset(dataset).findings
    assert [(finding.rule, finding.row, finding.field) for finding in findings] == [
        ("INVALID_NUMBER", 3, "power_kw")
    ]


def test_blank_csv_lines_do_not_shift_source_row_numbers() -> None:
    content = (
        ",".join(CANONICAL_COLUMNS)
        + "\n"
        + "PNL-A,Main Panel,panel,Plant 1,,,400,0,active\n"
        + "\n"
        + "MTR-001,Motor,motor,Plant 1,PNL-A,C-01,400,-1,active\n"
    ).encode()

    dataset = parse_dataset("assets.csv", content)

    assert dataset.row_numbers == [2, 4]
    assert dataset.records[1]["power_kw"] == -1


def test_literal_na_text_is_not_coerced_to_missing() -> None:
    content = (",".join(CANONICAL_COLUMNS) + "\n" + "PNL-A,NA,panel,N/A,,,400,0,active\n").encode()

    dataset = parse_dataset("assets.csv", content)

    assert dataset.records[0]["asset_name"] == "NA"
    assert dataset.records[0]["location"] == "N/A"


def test_exact_duplicate_xlsx_headers_are_rejected() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["asset_tag", "asset_tag"])
    sheet.append(["PNL-A", "PNL-B"])
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(DatasetError, match="Duplicate columns"):
        parse_dataset("assets.xlsx", output.getvalue())


def test_row_limit_is_enforced_before_validation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(ingest, "MAX_ROWS", 1)
    content = (
        ",".join(CANONICAL_COLUMNS)
        + "\n"
        + "PNL-A,Panel A,panel,Plant 1,,,400,0,active\n"
        + "PNL-B,Panel B,panel,Plant 1,,,400,0,active\n"
    ).encode()

    with pytest.raises(DatasetError, match="no more than 1 data rows"):
        parse_dataset("assets.csv", content)


def test_xlsx_row_limit_counts_non_empty_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(ingest, "MAX_ROWS", 1)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(CANONICAL_COLUMNS)
    sheet.append([None] * len(CANONICAL_COLUMNS))
    sheet.append(["PNL-A", "Main Panel", "panel", "Plant 1", None, None, 400, 0, "active"])
    output = BytesIO()
    workbook.save(output)

    dataset = parse_dataset("sparse.xlsx", output.getvalue())

    assert dataset.row_numbers == [3]
    assert len(dataset.records) == 1


@pytest.mark.parametrize(
    ("filename", "content", "message"),
    [
        ("assets.txt", b"anything", "Only .csv and .xlsx"),
        ("assets.csv", b"", "empty"),
        ("assets.csv", b"asset_tag\n", "data rows"),
        (
            "assets.csv",
            b"Asset Tag,asset_tag\nA,B\n",
            "Duplicate columns",
        ),
        (
            "assets.csv",
            b"asset_tag,asset_tag\nA,B\n",
            "Duplicate columns",
        ),
    ],
)
def test_malformed_datasets_are_rejected(filename: str, content: bytes, message: str) -> None:
    with pytest.raises(DatasetError, match=message):
        parse_dataset(filename, content)
