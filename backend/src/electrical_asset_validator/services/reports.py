from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Any, cast
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from ..models import ValidationRun
from .ingest import CANONICAL_COLUMNS

HEADER_FILL = PatternFill("solid", fgColor="123B5D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MAX_PDF_TEXT_CHARACTERS = 500


def _utc_isoformat(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    else:
        value = value.astimezone(UTC)
    return value.isoformat().replace("+00:00", "Z")


def _append_excel_row(sheet: Worksheet, values: list[Any] | tuple[Any, ...]) -> None:
    row_values = list(values)
    sheet.append(row_values)
    row_number = sheet.max_row
    for column_number, value in enumerate(row_values, start=1):
        if not isinstance(value, str) or not value.startswith(("=", "+", "-", "@")):
            continue
        cell = sheet.cell(row=row_number, column=column_number)
        # Preserve the exact source text while forcing Excel to display it as a
        # literal. Prefixing an apostrophe would exceed Excel's 32,767-character
        # cell limit for the longest accepted value and silently lose data.
        cell.data_type = "s"
        cell.quotePrefix = True


def _style_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    # Both report sheets are built from cell A1, so enumerating the columns
    # yields their real positions without reading the stub-optional Cell.column.
    for column_number, column_cells in enumerate(sheet.columns, start=1):
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_number)].width = min(
            max(width + 2, 10), 45
        )


def build_excel_report(validation: ValidationRun) -> bytes:
    workbook = Workbook()
    # The stubs mark .active as optional and possibly read-only; a Workbook
    # created in write mode always starts with one writable worksheet.
    summary = cast(Worksheet, workbook.active)
    summary.title = "Summary"
    _append_excel_row(summary, ["Metric", "Value"])
    for label, value in (
        ("Validation ID", validation.id),
        ("Source file", validation.filename),
        ("Created at", _utc_isoformat(validation.created_at)),
        ("Quality score", validation.quality_score),
        ("Total rows", validation.total_rows),
        ("Valid rows", validation.valid_rows),
        ("Issues", validation.issue_count),
        ("Returned issue details", validation.returned_issue_count),
        ("Issue details truncated", validation.issues_truncated),
        ("Errors", validation.error_count),
        ("Warnings", validation.warning_count),
        ("Information", validation.info_count),
    ):
        _append_excel_row(summary, [label, value])
    _style_sheet(summary)

    issues = workbook.create_sheet("Issues")
    _append_excel_row(
        issues,
        [
            "severity",
            "rule",
            "row",
            "asset_tag",
            "field",
            "message",
            "suggestion",
        ],
    )
    for issue in validation.issues:
        _append_excel_row(
            issues,
            [
                issue.severity,
                issue.rule,
                issue.row_number,
                issue.asset_tag,
                issue.field,
                issue.message,
                issue.suggestion,
            ],
        )
    _style_sheet(issues)

    data = workbook.create_sheet("Data")
    _append_excel_row(data, CANONICAL_COLUMNS)
    for record in validation.records:
        _append_excel_row(
            data,
            [record.get(column) for column in CANONICAL_COLUMNS],
        )
    _style_sheet(data)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _paragraph(value: Any, style: ParagraphStyle) -> Paragraph:
    raw_text = "" if value is None else str(value)
    if len(raw_text) > MAX_PDF_TEXT_CHARACTERS:
        raw_text = raw_text[: MAX_PDF_TEXT_CHARACTERS - 1] + "…"
    text = escape(raw_text)
    return Paragraph(text, style)


def build_pdf_report(validation: ValidationRun) -> bytes:
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"Validation report {validation.id}",
        author="Electrical Asset Validator",
    )
    styles = getSampleStyleSheet()
    body = ParagraphStyle(
        "ReportBody",
        parent=styles["BodyText"],
        fontSize=7,
        leading=9,
        alignment=TA_LEFT,
    )
    story: list[Any] = [
        Paragraph("Electrical Asset Validation Report", styles["Title"]),
        Paragraph(
            f"Source: {escape(validation.filename)} &nbsp;&nbsp; "
            f"Validation ID: {escape(validation.id)}",
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
    ]

    summary_data = [
        ["Quality score", "Rows", "Valid rows", "Errors", "Warnings", "Information"],
        [
            f"{validation.quality_score:.1f}",
            str(validation.total_rows),
            str(validation.valid_rows),
            str(validation.error_count),
            str(validation.warning_count),
            str(validation.info_count),
        ],
    ]
    summary_table = Table(summary_data, repeatRows=1)
    summary_table.setStyle(_table_style())
    story.extend([summary_table, Spacer(1, 6 * mm)])

    story.append(Paragraph("Findings", styles["Heading2"]))
    story.append(
        Paragraph(
            "Long values are truncated in this PDF for readability; "
            "the XLSX report preserves the accepted source values.",
            styles["BodyText"],
        )
    )
    story.append(Spacer(1, 2 * mm))
    if validation.issues:
        issue_data: list[list[Any]] = [
            ["Severity", "Rule", "Row", "Asset tag", "Field", "Message", "Suggestion"]
        ]
        for issue in validation.issues:
            issue_data.append(
                [
                    _paragraph(issue.severity, body),
                    _paragraph(issue.rule, body),
                    _paragraph(issue.row_number, body),
                    _paragraph(issue.asset_tag, body),
                    _paragraph(issue.field, body),
                    _paragraph(issue.message, body),
                    _paragraph(issue.suggestion, body),
                ]
            )
        issue_table = Table(
            issue_data,
            repeatRows=1,
            colWidths=[18 * mm, 35 * mm, 12 * mm, 24 * mm, 24 * mm, 95 * mm, 50 * mm],
        )
        issue_table.setStyle(_table_style())
        story.append(issue_table)
    else:
        story.append(Paragraph("No issues were found.", styles["BodyText"]))

    document.build(story)
    return output.getvalue()


def _table_style() -> TableStyle:
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123B5D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B8C4CE")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5F7")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
    )
