from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

import pandas as pd
from openpyxl import load_workbook

CANONICAL_COLUMNS = [
    "asset_tag",
    "asset_name",
    "asset_type",
    "location",
    "panel_tag",
    "circuit_ref",
    "voltage_v",
    "power_kw",
    "status",
]
NUMERIC_COLUMNS = {"voltage_v", "power_kw"}

# These limits bound work after the API's compressed upload-size check. They
# protect the synchronous rule engine, database, and JSON response from
# unexpectedly large tabular inputs and highly compressed XLSX workbooks.
MAX_ROWS = 50_000
MAX_COLUMNS = 64
MAX_CELL_CHARACTERS = 32_767
MAX_XLSX_ENTRIES = 2_000
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200
MAX_SCANNED_XLSX_ROWS = 250_000


class DatasetError(ValueError):
    """Raised when an uploaded register cannot be parsed safely."""


class MappingError(ValueError):
    """Raised when a column mapping cannot be applied to an upload."""


@dataclass(frozen=True)
class Dataset:
    records: list[dict[str, Any]]
    row_numbers: list[int]
    source_columns: list[str]
    present_columns: set[str]
    unexpected_columns: list[str]


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    header = str(value).strip().lower()
    header = re.sub(r"[\s-]+", "_", header)
    return re.sub(r"_+", "_", header)


def clean_scalar(value: Any) -> Any | None:
    if value is None:
        return None
    try:
        if bool(pd.isna(value)):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if hasattr(value, "isoformat") and not isinstance(value, (int, float, bool)):
        try:
            return value.isoformat()
        except (TypeError, ValueError):
            pass
    return value


def _normalize_numeric(value: Any | None) -> Any | None:
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return int(value) if isinstance(value, float) and value.is_integer() else value
    if not isinstance(value, str):
        return value
    try:
        parsed = float(value)
    except ValueError:
        return value
    if not pd.notna(parsed) or parsed in (float("inf"), float("-inf")):
        return value
    return int(parsed) if parsed.is_integer() else parsed


def _is_blank_row(row: list[Any] | tuple[Any, ...]) -> bool:
    return all(clean_scalar(value) is None for value in row)


def _check_cell_lengths(rows: list[list[Any]]) -> None:
    for row_number, row in enumerate(rows, start=1):
        for value in row:
            if isinstance(value, str) and len(value) > MAX_CELL_CHARACTERS:
                raise DatasetError(
                    "A cell exceeds the maximum supported length of "
                    f"{MAX_CELL_CHARACTERS:,} characters."
                )


def _read_csv(content: bytes) -> tuple[list[Any], list[list[Any]], list[int]]:
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise DatasetError("CSV files must use UTF-8 encoding.") from exc

    reader = csv.reader(StringIO(decoded, newline=""), strict=True)
    try:
        header = next(reader)
    except StopIteration as exc:
        raise DatasetError("The uploaded file does not contain a header row.") from exc
    except csv.Error as exc:
        raise DatasetError(f"The CSV header could not be read: {exc}.") from exc

    rows: list[list[Any]] = []
    row_numbers: list[int] = []
    try:
        while True:
            source_row = reader.line_num + 1
            row = next(reader)
            if _is_blank_row(row):
                continue
            if len(row) != len(header):
                raise DatasetError(
                    f"CSV row {source_row} has {len(row)} fields; "
                    f"the header has {len(header)}."
                )
            rows.append(row)
            row_numbers.append(source_row)
            if len(rows) > MAX_ROWS:
                raise DatasetError(
                    f"Files may contain no more than {MAX_ROWS:,} data rows."
                )
    except StopIteration:
        pass
    except csv.Error as exc:
        raise DatasetError(f"The CSV file could not be read: {exc}.") from exc

    return list(header), rows, row_numbers


def _preflight_xlsx(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
    except BadZipFile as exc:
        raise DatasetError("The XLSX file is not a valid workbook archive.") from exc

    if len(entries) > MAX_XLSX_ENTRIES:
        raise DatasetError(
            f"XLSX workbooks may contain no more than {MAX_XLSX_ENTRIES:,} files."
        )
    uncompressed_bytes = sum(entry.file_size for entry in entries)
    compressed_bytes = max(1, sum(entry.compress_size for entry in entries))
    if uncompressed_bytes > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise DatasetError(
            "The expanded XLSX workbook exceeds the "
            f"{MAX_XLSX_UNCOMPRESSED_BYTES // (1024 * 1024)} MB safety limit."
        )
    if (
        uncompressed_bytes > 5 * 1024 * 1024
        and uncompressed_bytes / compressed_bytes > MAX_XLSX_COMPRESSION_RATIO
    ):
        raise DatasetError("The XLSX workbook has an unsafe compression ratio.")


def _read_xlsx(content: bytes) -> tuple[list[Any], list[list[Any]], list[int]]:
    _preflight_xlsx(content)
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise DatasetError("The XLSX workbook could not be read.") from exc

    try:
        if not workbook.worksheets:
            raise DatasetError("The XLSX workbook does not contain a worksheet.")
        sheet = workbook.worksheets[0]
        if sheet.max_column > MAX_COLUMNS:
            raise DatasetError(
                f"Files may contain no more than {MAX_COLUMNS} columns."
            )
        if sheet.max_row > MAX_SCANNED_XLSX_ROWS:
            raise DatasetError(
                "The first XLSX worksheet extends beyond the supported "
                f"{MAX_SCANNED_XLSX_ROWS:,} source rows."
            )

        iterator = sheet.iter_rows(values_only=True)
        try:
            header_values = list(next(iterator))
        except StopIteration as exc:
            raise DatasetError(
                "The uploaded file does not contain a header row."
            ) from exc

        while header_values and clean_scalar(header_values[-1]) is None:
            header_values.pop()

        rows: list[list[Any]] = []
        row_numbers: list[int] = []
        for source_row, values in enumerate(iterator, start=2):
            row = list(values)
            if _is_blank_row(row):
                continue
            overflow = row[len(header_values) :]
            if any(clean_scalar(value) is not None for value in overflow):
                raise DatasetError(
                    f"XLSX row {source_row} contains data beyond the header columns."
                )
            row = row[: len(header_values)]
            if len(row) < len(header_values):
                row.extend([None] * (len(header_values) - len(row)))
            rows.append(row)
            row_numbers.append(source_row)
            if len(rows) > MAX_ROWS:
                raise DatasetError(
                    f"Files may contain no more than {MAX_ROWS:,} data rows."
                )
    finally:
        workbook.close()

    return header_values, rows, row_numbers


def _build_dataset(
    raw_headers: list[Any],
    rows: list[list[Any]],
    row_numbers: list[int],
    mapping: dict[str, str] | None = None,
) -> Dataset:
    if not raw_headers:
        raise DatasetError("The uploaded file does not contain a header row.")
    if len(raw_headers) > MAX_COLUMNS:
        raise DatasetError(f"Files may contain no more than {MAX_COLUMNS} columns.")
    if not rows:
        raise DatasetError("The uploaded file does not contain any data rows.")

    _check_cell_lengths([raw_headers, *rows])
    source_columns = ["" if value is None else str(value) for value in raw_headers]
    normalized_columns = [normalize_header(column) for column in raw_headers]
    if any(not column for column in normalized_columns):
        raise DatasetError("Column names cannot be blank.")
    duplicates = sorted(
        {
            column
            for column in normalized_columns
            if normalized_columns.count(column) > 1
        }
    )
    if duplicates:
        joined = ", ".join(duplicates)
        raise DatasetError(f"Duplicate columns after normalization: {joined}.")

    if mapping:
        missing_sources = sorted(set(mapping) - set(normalized_columns))
        if missing_sources:
            joined = ", ".join(missing_sources)
            raise MappingError(
                f"Mapped source headers are not present in the file: {joined}."
            )
        normalized_columns = [
            mapping.get(column, column) for column in normalized_columns
        ]
        mapped_duplicates = sorted(
            {
                column
                for column in normalized_columns
                if normalized_columns.count(column) > 1
            }
        )
        if mapped_duplicates:
            joined = ", ".join(mapped_duplicates)
            raise MappingError(
                f"Applying the mapping produces duplicate columns: {joined}."
            )

    frame = pd.DataFrame(rows, columns=normalized_columns, dtype=object)
    present_columns = set(normalized_columns) & set(CANONICAL_COLUMNS)
    unexpected_columns = [
        column for column in normalized_columns if column not in CANONICAL_COLUMNS
    ]

    records: list[dict[str, Any]] = []
    for raw_record in frame.to_dict(orient="records"):
        record: dict[str, Any] = {}
        for column in CANONICAL_COLUMNS:
            value = clean_scalar(raw_record.get(column))
            if column in NUMERIC_COLUMNS:
                value = _normalize_numeric(value)
            elif value is not None and not isinstance(value, str):
                value = str(value)
            record[column] = value
        records.append(record)

    return Dataset(
        records=records,
        row_numbers=row_numbers,
        source_columns=source_columns,
        present_columns=present_columns,
        unexpected_columns=unexpected_columns,
    )


def parse_column_mapping(raw: str) -> dict[str, str]:
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise MappingError(
            "The column mapping must be a valid JSON object of source "
            "headers to canonical fields."
        ) from exc
    if not isinstance(parsed, dict) or not all(
        isinstance(target, str) for target in parsed.values()
    ):
        raise MappingError(
            "The column mapping must be a JSON object of source headers "
            "to canonical fields."
        )

    mapping: dict[str, str] = {}
    for source, target in parsed.items():
        normalized_source = normalize_header(source)
        if not normalized_source:
            raise MappingError("Mapped source headers cannot be blank.")
        if target not in CANONICAL_COLUMNS:
            raise MappingError(
                f"'{target}' is not a canonical field. Canonical fields "
                "are: " + ", ".join(CANONICAL_COLUMNS) + "."
            )
        if normalized_source in mapping:
            raise MappingError(
                "Duplicate mapped source headers after normalization: "
                f"{normalized_source}."
            )
        mapping[normalized_source] = target

    targets = list(mapping.values())
    duplicate_targets = sorted(
        {target for target in targets if targets.count(target) > 1}
    )
    if duplicate_targets:
        joined = ", ".join(duplicate_targets)
        raise MappingError(
            f"Multiple source headers map to the same canonical field: {joined}."
        )
    return mapping


def parse_dataset(
    filename: str,
    content: bytes,
    mapping: dict[str, str] | None = None,
) -> Dataset:
    extension = Path(filename).suffix.lower()
    if extension not in {".csv", ".xlsx"}:
        raise DatasetError("Only .csv and .xlsx files are supported.")
    if not content:
        raise DatasetError("The uploaded file is empty.")

    if extension == ".csv":
        headers, rows, row_numbers = _read_csv(content)
    else:
        headers, rows, row_numbers = _read_xlsx(content)
    return _build_dataset(headers, rows, row_numbers, mapping)
